import pandas as pd
import os
from datetime import timedelta
from src.proyecto.config import RUTA_ARCHIVO_ASISTENCIA_DEFECTO, RUTA_SALIDA_ASISTENCIA_DEFECTO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

class AsistenciaService:

    def __init__(self, ruta_archivo: str | None = None, ruta_salida: str | None = None,fecha_desde: str | None = None, fecha_hasta: str | None = None):
        """
        Inicializa el servicio con rutas opcionales.
        Si no se proporcionan, usa las rutas por defecto.
        """
        self.ruta_archivo = ruta_archivo or RUTA_ARCHIVO_ASISTENCIA_DEFECTO
        self.ruta_salida = ruta_salida or RUTA_SALIDA_ASISTENCIA_DEFECTO
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
    # ==============================
    # CARGAR ARCHIVO
    # ==============================
    def cargar_archivo_excel(self) -> pd.DataFrame:
        if not os.path.exists(self.ruta_archivo):
            raise FileNotFoundError(
                f"El archivo no existe en la ruta: {self.ruta_archivo}"
            )

        try:
            df = pd.read_excel(self.ruta_archivo)
            return df
        except Exception as e:
            raise RuntimeError(f"Error al cargar el archivo: {e}")


    # ==============================
    # VALIDAR COLUMNAS
    # ==============================
    def validar_columnas(self, df: pd.DataFrame) -> None:
        columnas_requeridas = {"Codigo", "Fecha y hora"}

        if not columnas_requeridas.issubset(df.columns):
            raise ValueError(
                "El archivo no contiene las columnas requeridas: "
                "'Codigo' y 'Fecha y hora'"
            )


    # ==============================
    # PREPARAR FECHAS
    # ==============================
    def preparar_fechas(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            df["Fecha y hora"] = pd.to_datetime(df["Fecha y hora"])
        except Exception:
            raise ValueError(
                "La columna 'Fecha y hora' no tiene un formato válido."
            )

        if self.fecha_desde:
            fecha_desde_dt = pd.to_datetime(self.fecha_desde)
            df = df[df["Fecha y hora"] >= fecha_desde_dt]

        if self.fecha_hasta:
            fecha_hasta_dt = pd.to_datetime(self.fecha_hasta) + pd.Timedelta(days=1)
            df = df[df["Fecha y hora"] < fecha_hasta_dt]
        # ==============================
        # CREAR COLUMNAS AUXILIARES
        # ==============================
        df["año"] = df["Fecha y hora"].dt.year
        df["mes"] = df["Fecha y hora"].dt.month
        df["dia"] = df["Fecha y hora"].dt.day
        df["semana"] = df["Fecha y hora"].dt.isocalendar().week

        return df


    # ==============================
    # PROCESAR ASISTENCIA
    # ==============================
    def procesar_asistencia(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            df = df.sort_values(by=["Codigo", "Fecha y hora"])
            resultado = []

            for codigo, grupo in df.groupby("Codigo"):

                grupo = grupo.sort_values("Fecha y hora").reset_index(drop=True)
                i = 0

                while i < len(grupo) - 1:

                    hora_entrada = grupo.loc[i, "Fecha y hora"]
                    hora_salida = grupo.loc[i + 1, "Fecha y hora"]

                    diferencia_segundos = (hora_salida - hora_entrada).total_seconds()
                    diferencia_horas = diferencia_segundos / 3600

                    # 🟡 NUEVO: ignorar duplicados menores a 30 minutos
                    if diferencia_segundos < 1800:  # 30 * 60
                        i += 1
                        continue

                    # 🔴 Caso: posible falta de salida
                    if diferencia_horas > 13:
                        resultado.append({
                            "Código": codigo,
                            "Año": hora_entrada.year,
                            "Mes": hora_entrada.month,
                            "Semana": hora_entrada.isocalendar().week,
                            "Día": hora_entrada.day,
                            "Hora_entrada": hora_entrada.time(),
                            "Hora_salida": None,
                            "Horas_trabajadas": "00:00",
                            "Observación": "Posible falta de salida"
                        })
                        i += 1
                        continue

                    # 🔵 Caso válido
                    horas_td = timedelta(seconds=diferencia_segundos)

                    horas_formateadas = (
                        f"{int(horas_td.total_seconds() // 3600):02d}:"
                        f"{int((horas_td.total_seconds() % 3600) // 60):02d}"
                    )

                    observacion = (
                        "Turno válido"
                        if diferencia_horas >= 8
                        else "No cumplió 8 horas"
                    )

                    resultado.append({
                        "Código": codigo,
                        "Año": hora_entrada.year,
                        "Mes": hora_entrada.month,
                        "Semana": hora_entrada.isocalendar().week,
                        "Día": hora_entrada.day,
                        "Hora_entrada": hora_entrada.time(),
                        "Hora_salida": hora_salida.time(),
                        "Horas_trabajadas": horas_formateadas,
                        "Observación": observacion
                    })

                    i += 2

            return pd.DataFrame(resultado)

        except Exception as e:
            raise RuntimeError(f"Error procesando asistencia: {e}")

    def generar_formato_semanal(self, df, ruta_salida):
        try:

            if df.empty:
                raise ValueError("El DataFrame está vacío.")

            columnas_requeridas = [
                "Código", "Año", "Mes", "Semana",
                "Día", "Hora_entrada", "Hora_salida"
            ]

            for col in columnas_requeridas:
                if col not in df.columns:
                    raise ValueError(f"Falta la columna: {col}")

            wb = Workbook()
            wb.remove(wb.active)

            dias_semana = ["LUNES", "MARTES", "MIÉRCOLES",
                        "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]

            meses_es = {
                1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",
                5:"MAYO",6:"JUNIO",7:"JULIO",8:"AGOSTO",
                9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"
            }

            # 🔹 Agrupar por Año y Semana
            for (anio, semana), grupo_semana in df.groupby(["Año", "Semana"]):

                ws = wb.create_sheet(title=f"{anio}-S{semana}")

                mes = int(grupo_semana["Mes"].iloc[0])
                nombre_mes = meses_es.get(mes, "")

                # 🔹 Obtener lunes real de la semana ISO
                lunes_semana = datetime.fromisocalendar(int(anio), int(semana), 1)

                # 🔹 TÍTULO
                ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=16)
                ws["C1"] = f"SEMANA {semana} - {nombre_mes} {anio}"
                ws["C1"].alignment = Alignment(horizontal="center")
                ws["C1"].font = Font(bold=True, size=14)

                # 🔹 Encabezado OPERADOR
                ws.merge_cells("A3:A5")
                ws["A3"] = "OPERADOR"
                ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A3"].font = Font(bold=True)

                col = 2

                for i, dia_nombre in enumerate(dias_semana):

                    fecha_real = lunes_semana.replace() + \
                                (timedelta(days=i))

                    numero_dia = fecha_real.day

                    # Nombre del día
                    ws.merge_cells(start_row=3, start_column=col,
                                end_row=3, end_column=col+1)

                    ws.cell(row=3, column=col).value = dia_nombre
                    ws.cell(row=3, column=col).alignment = Alignment(horizontal="center")
                    ws.cell(row=3, column=col).font = Font(bold=True)

                    # Número del día
                    ws.merge_cells(start_row=4, start_column=col,
                                end_row=4, end_column=col+1)

                    ws.cell(row=4, column=col).value = numero_dia
                    ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")

                    # Entrada / Salida
                    ws.cell(row=5, column=col).value = "ENTRADA"
                    ws.cell(row=5, column=col+1).value = "SALIDA"

                    col += 2

                # 🔹 Insertar datos
                fila_excel = 6

                for codigo, grupo_operador in grupo_semana.groupby("Código"):

                    ws.cell(row=fila_excel, column=1).value = codigo
                    col = 2

                    for i in range(7):

                        fecha_real = lunes_semana + \
                                    timedelta(days=i)

                        dia_num = fecha_real.day

                        registro = grupo_operador[grupo_operador["Día"] == dia_num]

                        if not registro.empty:
                            hora_entrada = registro.iloc[0]["Hora_entrada"]
                            hora_salida = registro.iloc[0]["Hora_salida"]

                            if pd.notna(hora_entrada):
                                ws.cell(row=fila_excel, column=col).value = str(hora_entrada)

                            if pd.notna(hora_salida):
                                ws.cell(row=fila_excel, column=col+1).value = str(hora_salida)

                        col += 2

                    fila_excel += 1

                # Ajustar ancho columnas
                for i in range(1, 17):
                    ws.column_dimensions[get_column_letter(i)].width = 14

            wb.save(ruta_salida)

        except PermissionError:
            raise RuntimeError("No se pudo guardar el archivo. Cierra el Excel si está abierto.")

        except Exception as e:
            raise RuntimeError(f"Error generando formato semanal: {e}")

    # ==============================
    # ORDENAR
    # ==============================
    def ordenar_por_fecha(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            return df.sort_values(
                by=["Año", "Mes", "Día"]
            ).reset_index(drop=True)
        except Exception as e:
            raise RuntimeError(f"Error ordenando datos: {e}")

    # ==============================
    # GUARDAR
    # ==============================
    def guardar_asistencia(self, df: pd.DataFrame) -> str:
        try:
            nombre_archivo = "reporte_asistencia.xlsx"

            # Construir ruta final
            ruta_final = os.path.join(self.ruta_salida, nombre_archivo)

            if os.path.exists(ruta_final):
                df_existente = pd.read_excel(ruta_final)
                df_final = pd.concat([df_existente, df], ignore_index=True)
            else:
                df_final = df

            df_final.to_excel(ruta_final, index=False)

            return ruta_final

        except PermissionError:
            raise RuntimeError("No se pudo guardar el archivo. Cierra el Excel si está abierto.")
        
        except Exception as e:
            raise RuntimeError(f"Error guardando archivo: {e}")
    
    def ejecutar_proceso_dataframe(self) -> pd.DataFrame:
        """
        Ejecuta el proceso completo pero retorna el DataFrame resultante
        en lugar de guardarlo. Útil para pruebas unitarias.
        """
        try:
            df = self.cargar_archivo_excel()
            self.validar_columnas(df)
            df = self.preparar_fechas(df)
            df_procesado = self.procesar_asistencia(df)
            df_ordenado = self.ordenar_por_fecha(df_procesado)
            return df_ordenado

        except PermissionError:
            raise RuntimeError("No se pudo guardar el archivo. Cierra el Excel si está abierto.")
        
        except Exception as e:
            raise RuntimeError(f"Error en el proceso completo: {e}")

    # ==============================
    # PROCESO COMPLETO
    # ==============================
    def ejecutar_proceso_completo(self) -> str:
        """
        Ejecuta todo el flujo:
        - Cargar archivo
        - Validar columnas
        - Preparar fechas
        - Procesar asistencia
        - Ordenar resultados
        - Guardar archivo
        
        Retorna la ruta donde se guardó el archivo.
        """
        try:
            df_ordenado = self.ejecutar_proceso_dataframe()
            self.generar_formato_semanal(df_ordenado, os.path.join(self.ruta_salida, "reporte_semanal.xlsx"))
            ruta_guardada = self.guardar_asistencia(df_ordenado)
            return ruta_guardada

        except Exception as e:
            raise RuntimeError(f"Error en el proceso completo: {e}")